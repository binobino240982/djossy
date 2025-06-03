from django.shortcuts import render, redirect, get_object_or_404
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from django.db.models import Count
from django.http import HttpResponse
from django.contrib import messages

import csv
import openpyxl
from openpyxl.utils import get_column_letter

from .models import Candidature, DemandePersonnel, Profil
from .forms import (
    FormulaireRecrutement,
    CandidatureForm,
    DemandePersonnelForm,
)

# Vérifie si l'utilisateur est admin/staff
def est_admin(user):
    return user.is_staff

# Dashboard résumé (stats globales) - réservé staff/admin
@staff_member_required
def admin_dashboard(request):
    stats_candidatures = Candidature.objects.values('statut').annotate(total=Count('id'))
    stats_demandes = DemandePersonnel.objects.values('statut').annotate(total=Count('id'))

    context = {
        'stats_candidatures': stats_candidatures,
        'stats_demandes': stats_demandes,
        'total_candidatures': Candidature.objects.count(),
        'total_demandes': DemandePersonnel.objects.count(),
    }
    return render(request, 'recrutement/admin_dashboard.html', context)

# Dashboard détaillé (autre version) - accès restreint
@login_required
@user_passes_test(est_admin)
def dashboard_admin(request):
    context = {
        'total_candidatures': Candidature.objects.count(),
        'nouvelles': Candidature.objects.filter(statut='nouvelle').count(),
        'en_cours': Candidature.objects.filter(statut='en_cours').count(),
        'accepte': Candidature.objects.filter(statut='acceptee').count(),
        'rejete': Candidature.objects.filter(statut='rejetee').count(),
        'demandes': DemandePersonnel.objects.count(),
        'chart_data': {
            'labels': ['Nouvelles', 'En cours', 'Acceptées', 'Rejetées'],
            'values': [
                Candidature.objects.filter(statut='nouvelle').count(),
                Candidature.objects.filter(statut='en_cours').count(),
                Candidature.objects.filter(statut='acceptee').count(),
                Candidature.objects.filter(statut='rejetee').count()
            ]
        }
    }
    return render(request, 'gestion_admin/dashboard.html', context)

# Accueil du dashboard public
def accueil_dashboard(request):
    return render(request, 'dashboard.html')
    
def personnel_accueil(request):
    return render(request, 'personnel/accueil.html')

# Liste des candidatures (filtrage possible par statut) - accès admin
@login_required
@user_passes_test(est_admin)
def liste_candidatures(request):
    statut = request.GET.get('statut')
    candidatures_list = (
        Candidature.objects.filter(statut=statut).order_by('-date_postulation')
        if statut else
        Candidature.objects.all().order_by('-date_postulation')
    )

    paginator = Paginator(candidatures_list, 10)  # 10 candidatures par page
    page = request.GET.get('page')
    candidatures = paginator.get_page(page)
    return render(request, 'gestion_admin/liste_candidatures.html', {'candidatures': candidatures})

# Liste des demandes de personnel - accès admin
@login_required
@user_passes_test(est_admin)
def liste_demandes(request):
    demandes = DemandePersonnel.objects.all().order_by('-date_demande')
    return render(request, 'gestion_admin/liste_demandes.html', {'demandes': demandes})

# Affiche tous les profils (public)
def liste_profils(request):
    profils = Profil.objects.all()
    return render(request, 'personnel/liste_profils.html', {'profils': profils})

# Affiche le détail d’un profil (public)
def detail_profil(request, profil_id):
    profil = get_object_or_404(Profil, id=profil_id)
    return render(request, 'personnel/detail_profil.html', {'profil': profil})

# Formulaire de recrutement pour un profil spécifique
def recruter(request, profil_id):
    profil = get_object_or_404(Profil, id=profil_id)
    if request.method == 'POST':
        form = FormulaireRecrutement(request.POST, request.FILES)
        if form.is_valid():
            candidature = form.save(commit=False)
            candidature.profil = profil
            candidature.save()

            # Envoi mail notification
            sujet = f"Nouvelle candidature pour {profil.nom}"
            message = (
                f"{candidature.nom} a postulé pour {profil.nom}.\n\n"
                f"Email: {candidature.email}\n\nMessage:\n{candidature.message}"
            )
            destinataire = ['tonemail@tonsite.com']  # À changer par ton email réel
            send_mail(sujet, message, settings.DEFAULT_FROM_EMAIL, destinataire)

            messages.success(request, "Votre candidature a été envoyée avec succès.")
            return redirect('personnel:detail_profil', profil_id=profil.id)
    else:
        form = FormulaireRecrutement()
    return render(request, 'personnel/recruter.html', {'form': form, 'profil': profil})

# Candidature libre
def postuler(request):
    if request.method == 'POST':
        form = CandidatureForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Votre candidature a bien été soumise.")
            return redirect('personnel:confirmation_candidature')
        else:
            messages.error(request, "Erreur dans le formulaire. Vérifiez les champs.")
    else:
        form = CandidatureForm()
    return render(request, 'recrutement/postuler.html', {'form': form})

# Page de confirmation de candidature
def confirmation_candidature(request):
    return render(request, 'candidature/confirmation_candidature.html')

# Formulaire demande pour employeurs
def soumettre_demande_personnel(request):
    if request.method == 'POST':
        form = DemandePersonnelForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Votre demande a été envoyée avec succès.")
            return redirect('personnel:liste_profils')
    else:
        form = DemandePersonnelForm()
    return render(request, 'personnel/soumettre_demande.html', {'form': form})

# Remerciement
def merci(request):
    return render(request, 'personnel/merci.html')

# Export CSV des candidatures - accès admin
@login_required
@user_passes_test(est_admin)
def exporter_candidatures_csv(request):
    candidatures = Candidature.objects.all()
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="candidatures.csv"'
    writer = csv.writer(response)
    writer.writerow(['Nom', 'Email', 'Statut', 'Date de postulation'])

    for c in candidatures:
        writer.writerow([c.nom, c.email, c.statut, c.date_postulation])
    return response

# Export Excel des candidatures - accès admin
@login_required
@user_passes_test(est_admin)
def export_candidatures_excel(request):
    candidatures = Candidature.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidatures"
    headers = ['Nom', 'Email', 'Statut', 'Date de postulation']
    ws.append(headers)

    for c in candidatures:
        ws.append([
            c.nom, c.email, c.statut,
            c.date_postulation.strftime('%Y-%m-%d %H:%M') if c.date_postulation else ''
        ])

    for col_num, column_title in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = max(len(column_title), 15)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=candidatures.xlsx'
    wb.save(response)
    return response
